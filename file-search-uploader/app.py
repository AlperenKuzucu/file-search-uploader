# app.py
from fastapi import FastAPI, Request, UploadFile, File
from fastapi.responses import RedirectResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

import os, sqlite3, time, mimetypes, re
from typing import List, Tuple, Optional
from datetime import datetime

import pwd
from docx import Document as DocxDocument
from openpyxl import load_workbook
from PyPDF2 import PdfReader


try:
    import xlrd  
    XLRD_AVAILABLE = True
except Exception:
    xlrd = None  
    XLRD_AVAILABLE = False



APP_DIR    = os.path.abspath(os.path.dirname(__file__))
UPLOAD_DIR = os.path.join(APP_DIR, "uploads")
DB_PATH    = os.path.join(APP_DIR, "files.db")
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = FastAPI(title="Çoklu Dosya Yükleme & Arama")
app.mount("/static", StaticFiles(directory=os.path.join(APP_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(APP_DIR, "templates"))



def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = db()
    conn.execute("""
    CREATE TABLE IF NOT EXISTS files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        filename    TEXT,
        path        TEXT,
        filetype    TEXT,
        author      TEXT,
        created     TEXT,
        owner       TEXT,
        size_bytes  INTEGER,
        uploaded_at TEXT,
        content     TEXT
    )
    """)
    conn.commit(); conn.close()

init_db()



def file_owner_username(path: str) -> Optional[str]:
    try:
        st = os.stat(path)
        return pwd.getpwuid(st.st_uid).pw_name
    except Exception:
        return None

def clean_text(s: str) -> str:
    if not s: return ""
    return " ".join(str(s).split())

def sniff_type(filename: str) -> str:
    ext = os.path.splitext(filename)[1].lower()
    if ext in [".docx"]: return "docx"
    if ext in [".xlsx"]: return "xlsx"
    if ext in [".xls"] : return "xls"
    if ext in [".pdf"] : return "pdf"
    if ext in [".txt"] : return "txt"
    mt = (mimetypes.guess_type(filename)[0] or "").lower()
    if "pdf"   in mt: return "pdf"
    if "word"  in mt: return "docx"
    if "excel" in mt: return "xlsx"
    if "text"  in mt: return "txt"
    return ext.strip(".") or "unknown"

def extract_docx(path: str) -> Tuple[str, dict]:
    doc = DocxDocument(path)
    content = "\n".join([p.text for p in doc.paragraphs if p.text and p.text.strip()])
    props = doc.core_properties
    meta = {
        "author": props.author or None,
        "created": props.created.isoformat() if props.created else None,
    }
    return content, meta

def extract_xlsx(path: str) -> Tuple[str, dict]:
    wb = load_workbook(path, data_only=True)
    parts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c not in (None, "")]
            if cells: parts.append(" ".join(cells))
    content = "\n".join(parts)
    meta = {
        "author": wb.properties.creator or None,
        "created": wb.properties.created.isoformat() if wb.properties.created else None,
    }
    return content, meta

def extract_xls(path: str) -> Tuple[str, dict]:
    if not XLRD_AVAILABLE:
        return "", {"author": None, "created": None}
    book = xlrd.open_workbook(path)  
    parts = []
    for sh in book.sheets():
        for r in range(sh.nrows):
            row = [str(sh.cell_value(r, c)) for c in range(sh.ncols) if sh.cell_value(r, c) not in (None, "")]
            if row: parts.append(" ".join(row))
    return "\n".join(parts), {"author": None, "created": None}

def extract_pdf(path: str) -> Tuple[str, dict]:
    reader = PdfReader(path)
    content = "\n".join([(p.extract_text() or "") for p in reader.pages])
    meta_raw = reader.metadata or {}
    created = None
    raw = meta_raw.get("/CreationDate") or meta_raw.get("/ModDate")
    if raw:
        digits = "".join([c for c in raw if c.isdigit()])
        if len(digits) >= 14:
            try:
                created = datetime.strptime(digits[:14], "%Y%m%d%H%M%S").isoformat()
            except Exception:
                created = None
    author = meta_raw.get("/Author", None)
    return content, {"author": author, "created": created}

def extract_txt(path: str) -> Tuple[str, dict]:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read(), {"author": None, "created": None}

def extract_content_meta(path: str, ftype: str) -> Tuple[str, dict]:
    if ftype == "docx": return extract_docx(path)
    if ftype == "xlsx": return extract_xlsx(path)
    if ftype == "xls" : return extract_xls(path)
    if ftype == "pdf" : return extract_pdf(path)
    if ftype == "txt" : return extract_txt(path)
    return "", {"author": None, "created": None}

def save_record(filename: str, path: str, filetype: str, content: str,
                author: Optional[str], created: Optional[str]):
    conn = db()
    conn.execute("""
        INSERT INTO files (filename, path, filetype, author, created, owner, size_bytes, uploaded_at, content)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        filename, path, filetype,
        author, created,
        file_owner_username(path),
        os.path.getsize(path),
        datetime.now().isoformat(),
        clean_text(content),
    ))
    conn.commit(); conn.close()



def tr_fold(s: str) -> str:
    """İ/ı özelinde Türkçe casefold; uzun ß gibi genişleten yok, uzunluk korunur."""
    if not s:
        return ""
    s = s.replace("İ", "i").replace("I", "ı")
    return s.casefold()



def scope_texts(row, scope: str):
    if scope == "name":
        return [row["filename"] or ""]
    if scope == "content":
        return [row["content"] or ""]
    return [row["filename"] or "", row["content"] or ""]

def search_content(term: str, mode: str, scope: str):
    term = (term or "").strip()
    if not term:
        return []

    ft = tr_fold(term)

    
    if mode == "startswith":
        pat = re.compile(rf"(?<!\w){re.escape(ft)}", flags=re.UNICODE)
    elif mode == "endswith":
        pat = re.compile(rf"{re.escape(ft)}(?!\w)", flags=re.UNICODE)
    else:  
        pat = re.compile(rf"{re.escape(ft)}", flags=re.UNICODE)

    conn = db()
    try:
        
        rows = conn.execute("SELECT * FROM files ORDER BY uploaded_at DESC").fetchall()
    finally:
        conn.close()

    matched = []
    for r in rows:
        texts = scope_texts(r, scope)
        for t in texts:
            if not t:
                continue
            if pat.search(tr_fold(t)):
                matched.append(r)
                break
    return matched



def highlight_snippet(text: str, term: str, mode: str = "contains", pad: int = 60) -> str:
    """
    İçerikte ilk eşleşmeyi (Türkçe katlamayla) <mark> ile vurgular.
    startswith = kelime başı, endswith = kelime sonu.
    """
    s = text or ""
    t = (term or "").strip()
    if not s or not t:
        return ""

    fs = tr_fold(s)
    ft = tr_fold(t)

    if mode == "startswith":
        pat = re.compile(rf"(?<!\w){re.escape(ft)}", flags=re.UNICODE)
    elif mode == "endswith":
        pat = re.compile(rf"{re.escape(ft)}(?!\w)", flags=re.UNICODE)
    else:
        pat = re.compile(rf"{re.escape(ft)}", flags=re.UNICODE)

    m = pat.search(fs)
    if not m:
        return ""

    i, j = m.span()          

    start = max(0, i - pad)
    end   = min(len(s), j + pad)
    window = s[start:end].replace("\n", " ")
    i2, j2 = i - start, j - start

    before = window[:i2]
    hit    = window[i2:j2]
    after  = window[j2:]

    if len(before) > pad: before = "…" + before[-pad:]
    if len(after)  > pad: after  = after[:pad] + "…"
    return f"{before}<mark>{hit}</mark>{after}"

def make_highlighter(mode: str):
    def _hl(text: str, term: str) -> str:
        return highlight_snippet(text, term, mode=mode)
    return _hl



def human_size(n: int) -> str:
    try: n = int(n)
    except Exception: return ""
    for u in ["B","KB","MB","GB","TB"]:
        if n < 1024: return f"{n:.0f} {u}"
        n /= 1024
    return f"{n:.0f} PB"

def fmt_dt(iso_str: Optional[str]) -> str:
    if not iso_str: return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z","+00:00"))
        return dt.astimezone().strftime("%Y-%m-%d %H:%M")
    except Exception:
        return iso_str

templates.env.filters["human"] = human_size
templates.env.filters["dt"]    = fmt_dt



@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    conn = db()
    rows = conn.execute("""
        SELECT id, filename, filetype, author, created, owner, size_bytes, uploaded_at
        FROM files
        ORDER BY uploaded_at DESC
        LIMIT 100
    """).fetchall()
    conn.close()
    return templates.TemplateResponse("index.html", {"request": request, "files": rows})

@app.post("/upload")
async def upload(files: List[UploadFile] = File(...)):
    for uf in files:
        name = os.path.basename((uf.filename or "").strip())
        if not name:
            continue

        dest = os.path.join(UPLOAD_DIR, f"{int(time.time()*1000)}_{name}")
        with open(dest, "wb") as out:
            out.write(await uf.read())

        ftype = sniff_type(name)

        
        if os.path.getsize(dest) == 0:
            content, meta = "", {"author": None, "created": None}
            save_record(name, dest, ftype, content, meta.get("author"), meta.get("created"))
            continue

        
        if ftype == "unknown" and getattr(uf, "content_type", None):
            ct = uf.content_type.lower()
            if "plain" in ct: ftype = "txt"
            elif "pdf" in ct: ftype = "pdf"
            elif "word" in ct: ftype = "docx"
            elif "excel" in ct or "spreadsheet" in ct: ftype = "xlsx"

        content, meta = extract_content_meta(dest, ftype)
        save_record(
            name, dest, ftype, content,
            meta.get("author") or "Bilinmiyor",
            meta.get("created") or "Yok"
        )

    return RedirectResponse("/", status_code=303)

@app.get("/search", response_class=HTMLResponse)
def search(request: Request, term: str, mode: str = "contains", scope: str = "both"):
    
    if mode not in {"contains", "startswith", "endswith"}:
        mode = "contains"
    rows = search_content(term, mode, scope)
    return templates.TemplateResponse(
        "search.html",
        {
            "request": request,
            "rows": rows,
            "term": term,
            "mode": mode,
            "scope": scope,
            "highlight": make_highlighter(mode),
            "snippet": lambda txt, t, m=None: highlight_snippet(txt, t, mode if m is None else m),
        },
    )